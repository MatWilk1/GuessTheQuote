"""
Guess The Quote or Emilkowe Cytaty
Created by Mateusz Wilk

Quiz game where player guesses quotes from books or famous people.
Program uses excel file as a base of quotations and best scores (three different sheets).
"""


import random
import time
import openpyxl
import os


# Show menu
def menu():
    os.system("cls")
    print('\nWybierz pozycję z menu i naciśnij Enter \n')
    print('1. Rozpocznij grę')
    print('2. Zakończ grę')
    print('3. Pokaż najlepsze wyniki')
    print('4. Informacje o grze \n')
    menu1 = input()
    if menu1 == '1':
        os.system("cls")
        game()
    elif menu1 == '2':
        os.system("cls")
        print('\nDo zobaczenia ' + player + ' :)')
        time.sleep(2)
        exit()
    elif menu1 == '3':
        os.system("cls")
        best_scores()
    elif menu1 == '4':
        os.system("cls")
        print('\nEmilkowe Cytaty v1.0')
        print('\nCreated by Mateusz \'Wilq\' Wilk')
        input('\nNaciśnij Enter aby wrócić do menu')
        menu()
    else:
        menu()


# Game start. Choose game type. Only '1' or '2' allow to start.
def game():
    os.system("cls")
    print('\nWybierz kategorię:'
          '\n1. Fragmenty książek i cytaty znanych ludzi.'
          '\n2. Fragmenty książek.')

    category_1 = ''
    while category_1 not in ('1', '2'):
        category_1 = input()
        if category_1 not in ('1', '2'):
            print('Naciśnij prawidłowy klawisz')

    # Open excel and assign two sheets to variables
    file = openpyxl.load_workbook('QuotesBase.xlsx')
    sheet_b = file['books']
    sheet_p = file['people']

    score = 0

    if category_1 == "1":
        print('\nKategoria: Fragmenty książek i cytaty znanych ludzi.')
        for i in range(1, 6):
            category_2 = random.choice(['books', 'people'])
            os.system("cls")
            print('\nPytanie nr', i, '\n')
            if category_2 == "books":
                if question(sheet_b, score, 'books') == 1:
                    score += 1
                    correct_answer(score)

            elif category_2 == "people":
                if question(sheet_p, score, 'people') == 1:
                    score += 1
                    correct_answer(score)

            else:
                print("To jest błąd :(")

        print('\nTo już koniec gry')
        save_score(score)
        input('\nNaciśnij Enter aby wrócić do menu')
        menu()

    elif category_1 == "2":
        print('\nKategoria: Fragmenty książek')
        for i in range(1, 6):
            os.system("cls")
            print('\nPytanie nr', i, '\n')
            if question(sheet_b, score, 'books') == 1:
                score += 1
                correct_answer(score)

        print('\nTo już koniec gry')
        save_score(score)
        input('\nNaciśnij Enter aby wrócić do menu')
        menu()


def question(sheet_name, score, qestion_type):
    # Draw 4 quotes (from books or people). The first is the one to guess (and it is print)
    no = random.sample(range(1, sheet_name.max_row), 4)

    if qestion_type == 'books':
        print("Z jakiej książki pochodzi ten fragment?")
    elif qestion_type == 'people':
        print('Kto to powiedział?')

    print(sheet_name['B{}'.format(no[0])].value, '\n')

    # Shuffle drawn books titles and print possible answers
    no_shuffled = random.sample(no, 4)
    print('1.', sheet_name['A{}'.format(no_shuffled[0])].value)
    print('2.', sheet_name['A{}'.format(no_shuffled[1])].value)
    print('3.', sheet_name['A{}'.format(no_shuffled[2])].value)
    print('4.', sheet_name['A{}'.format(no_shuffled[3])].value)

    answer = ''
    while answer not in ('1', '2', '3', '4'):
        answer = input()
        if answer not in ('1', '2', '3', '4'):
            print('Naciśnij prawidłowy klawisz')

    # Check if answer is equal to first drawn book (first is the correct one)
    if answer == "1" and no_shuffled[0] == no[0]:
        return 1
    elif answer == "2" and no_shuffled[1] == no[0]:
        return 1
    elif answer == "3" and no_shuffled[2] == no[0]:
        return 1
    elif answer == "4" and no_shuffled[3] == no[0]:
        return 1
    else:
        print('Zła odpowiedź :(')
        print('Prawidłowa to: ', sheet_name['A{}'.format(no[0])].value, '\n')
        print('Twój Wynik =', score, 'pkt')
        input('\nNaciśnij Enter aby przejść dalej')
        return 0


# Show 5 the best scores from excel file ('scores' sheet)
def best_scores():
    file = openpyxl.load_workbook('QuotesBase.xlsx')
    sheet_s = file['scores']
    print('\nNajlepsze wyniki:')
    for i in range(1, 6):
        print(sheet_s['A{}'.format(i)].value)
    input('\nNaciśnij Enter aby wrócić do menu')
    menu()


# Save the best scores to excel file ('scores' sheet)
def save_score(score):
    file = openpyxl.load_workbook('QuotesBase.xlsx')
    sheet_s = file['scores']
    score = str(score)
    score_excell = (score + ' ' + '- ' + player)

    if sheet_s['A1'].value is None:
        for i in range(1, 6):
            i = str(i)
            sheet_s['A' + i] = '0'

    scores_tab = [sheet_s['A1'].value, sheet_s['A2'].value, sheet_s['A3'].value, sheet_s['A4'].value,
                  sheet_s['A5'].value, score_excell]

    scores_tab_sort = sorted(scores_tab, reverse=True)

    for i in range(1, 6):
        j = i
        i = str(i)
        sheet_s['A' + i] = scores_tab_sort[j-1]

    # Congrats if score is equel or greater than the best
    if score_excell == scores_tab_sort[0]:
        print('Gratulacje, to jest najlepszy wynik :)')

    file.save('QuotesBase.xlsx')


# Show one of few congrats when correct answer and show the score
def correct_answer(score):
    print(random.choice(['Super, zgadza się :)', 'Gratulacje, dobra odpowiedź :)', 'Brawo, poprawna odpowiedź :)']))
    print('\nTwój Wynik =', score, 'pkt')
    input('\nNaciśnij Enter aby przejść dalej')


print("""
 ____   __    __   _   _       _   __   ___     _      _   ____
|  __| |  \\  /  | | | | |     | | / /  /   \\   | |    | | |  __|
| |__  | |\\\//| | | | | |     | |/ /  /  _  \\  | |    | | | |__
|  __| | | \\/ | | | | | |     |   /  |  |_|  | | | /\\ | | |  __|
| |__  | |    | | | | | |___  | |\\ \\  \\     /  | |//\\\\| | | |__
|____| |_|    |_| |_| |_____| |_| \\_\\  \\___/   |__/  \\__| |____|

 ____   __    __  _______      ____      _______  __    __
|  __|  \\ \\  / / |__   __|    / __ \\    |__   __| \\ \\  / /
| |      \\ \\/ /     | |      / /__\\ \\      | |     \\ \\/ /
| |       \\  /      | |     / _____  \\     | |      \\  / 
| |__     |  |      | |    / /      \\ \\    | |      |  |
|____|    |__|      |_|   /_/        \\_\\   |_|      |__|

""")


time.sleep(2)
print('\nWitaj w grze Emilkowe Cytaty :) \n')

# Take a player name
player = ''
while player == '':
    player = input('Podaj swoje imię ')

    if player is '':
        print('Nic nie wpisałeś ')
    else:
        os.system("cls")
        print('\nCześć', player, ':)')

time.sleep(2)

menu()
